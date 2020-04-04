import openpyxl
from openpyxl.styles import PatternFill

wbkName = 'COLL.xlsx'
wbk = openpyxl.load_workbook(wbkName)
wks1 = wbk['Sheet1']
wks2 = wbk['Sheet2']

modDict = {}
modDict[4] = ""
def condense(x):
    #assume x % 6 == 4
    orig = x
    if x in modDict:
        return modDict[x]
    if x % 24 == 4:
        cycle = "B"
    elif x % 24 == 10:
        cycle = "C"
    elif x % 24 == 16:
        cycle = "A"
    elif x % 24 == 22:
        cycle = "D"
    else:
        print("error")
    x //= 2
    if x % 6 == 5:
        x = 3 * x + 1
    else:
        x //= 2
        if x % 6 != 4:
            x = 3 * x + 1
    if x in modDict:
        new = cycle + modDict[x]
    else:
        new = cycle + condense(x)
    modDict[orig] = new
    return new

r = 1
for i in range(4, 2000001, 6): #(X)change second parameter to add more rows
    p = condense(i)
    wks1.cell(row=r, column=1).value = i
    wks1.cell(row=r, column=2).value = p
    wks1.cell(row=r, column=3).value = len(p)
    r += 1

r = 1
for i in range(4, 2000001, 6): #(Y)change second parameter to add more rows, must have Y <= X
    p = modDict[i]
    last = 4
    mod = 6
    path = ""
    while len(path) < len(p):
        mod *= 4
        new = i % mod
        if new == last:
            path += "1"
        elif new == last + (mod/4):
            path += "2"
        elif new == last + (mod/2):
            path += "3"
        elif new == last + (3*mod/4):
            path += "4"
        else:
            print("error")
        last = new
    wks1.cell(row=r, column=4).value = path
    r += 1

def genNewList(old, increment):
    result = []
    for i in old:
        result.append(i)
        result.append(i + increment)
        result.append(i + 2 * increment)
        result.append(i + 3 * increment)
    return result

colsToDo = 4 #(Z)change to generate more columns, must have 4^Z <= X
modLists = []
modLists.append([4,10,16,22])
for i in range(1, colsToDo):
    modLists.append(genNewList(modLists[i - 1], 24 * 4**(i - 1)))
for i in range(0, colsToDo):
    r = 1
    for j in modLists[i]:
        wks2.cell(row=r, column=2 * i + 1).value = j
        if len(modDict[j]) < i + 1:
            wks2.cell(row=r, column=i * 2 + 2).value = "B"
        else:
            wks2.cell(row=r, column=i * 2 + 2).value = modDict[j][i]
        if len(modDict[j]) == i + 1:
            wks2.cell(row=r, column=2 * i + 1).fill = PatternFill(patternType='solid', fgColor='FFFF00')
            wks2.cell(row=r, column=2 * i + 2).fill = PatternFill(patternType='solid', fgColor='FFFF00')
        r += (4**(len(modLists) - i - 1))

wbk.save(wbkName)
wbk.close